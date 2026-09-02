"""Pure reader for .xlsx / .xlsm workbooks (openpyxl), kept import-light and
unit-testable. The handlers in main.py wrap these into JSON actions.
"""
from __future__ import annotations

from pathlib import Path


def _load(path: str, data_only: bool = True):
    from openpyxl import load_workbook
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"file not found: {path}")
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(f"not an .xlsx/.xlsm file: {path}")
    return load_workbook(str(p), data_only=data_only)


def list_sheets(path: str) -> list[dict]:
    """Return [{title, rows, columns}] for every worksheet."""
    wb = _load(path)
    out = []
    for ws in wb.worksheets:
        out.append({
            "title": ws.title,
            "rows": ws.max_row,
            "columns": ws.max_column,
        })
    return out


def _trim_row(row: tuple) -> list:
    """Drop trailing empty cells; stringify everything (None -> '')."""
    vals = list(row)
    while vals and (vals[-1] is None or str(vals[-1]).strip() == ""):
        vals.pop()
    return ["" if v is None else v for v in vals]


def _is_empty(vals: list) -> bool:
    return all(v is None or str(v).strip() == "" for v in vals)


def read_sheet(path: str, sheet: str | None = None, *,
               max_rows: int | None = None, skip_empty: bool = True) -> dict:
    """Return {sheet, rows: [[...]], truncated} for one worksheet.

    sheet=None reads the first/active worksheet. Cell values are JSON-safe
    (numbers/strings/bools; dates become their string form via openpyxl)."""
    wb = _load(path)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    rows: list[list] = []
    truncated = False
    for r in ws.iter_rows(values_only=True):
        vals = _trim_row(r)
        if skip_empty and _is_empty(vals):
            continue
        rows.append([_json_safe(v) for v in vals])
        if max_rows is not None and len(rows) >= max_rows:
            truncated = True
            break
    return {"sheet": ws.title, "rows": rows, "truncated": truncated}


def _json_safe(v):
    import datetime
    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        return v.isoformat()
    return v


def dump_text(path: str, *, max_rows: int | None = None) -> str:
    """Whole workbook as readable pipe-delimited text (one row per line,
    empty rows skipped), with sheet headers. Mirrors the original ad-hoc dump."""
    wb = _load(path)
    lines = [f"### FILE: {path}", f"### SHEETS: {wb.sheetnames}", ""]
    for ws in wb.worksheets:
        lines.append(f"\n===== SHEET: {ws.title}  (dims {ws.dimensions}) =====")
        shown = 0
        for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = _trim_row(r)
            if _is_empty(vals):
                continue
            lines.append(f"r{i}: " + " | ".join(str(_json_safe(v)) for v in vals))
            shown += 1
            if max_rows is not None and shown >= max_rows:
                lines.append(f"... [truncated at {max_rows} non-empty rows]")
                break
        if shown == 0:
            lines.append("(empty)")
    return "\n".join(lines)
